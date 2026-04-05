"""
sheets/views.py
===============
All HTTP views for LinkSheet.

View organisation:
  1. Public / auth  — home, google_login, google_callback, logout_view
  2. Dashboard       — dashboard
  3. Sheet CRUD      — create_google_sheet, delete_sheet, download_sheet
  4. Row CRUD        — add_row, update_row, delete_row
  5. Grid data       — sheet_grid_data
  6. Membership      — join_sheet, add_collaborator, remove_collaborator
  7. Sync            — trigger_sync
  8. Activity        — activity_page
  9. AJAX helpers    — get_created_sheets
"""
import csv
import json
import logging
from io import BytesIO

import google.auth.transport.requests
import openpyxl
import requests
from django.contrib import messages
from django.contrib.auth import get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Max, Prefetch, Q
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl.styles import Alignment, Font, PatternFill

from sheets.models import ActivityLog, GoogleCredentials, Sheet, SheetColumn, SheetMember, SheetRow, SheetSyncEvent
from sheets.permissions import (
    can_access_sheet, can_delete_row, can_download_sheet, can_manage_sheet,
    can_modify_row, can_see_all_rows, can_trigger_sync,
    get_role, get_visible_rows,
    is_collaborator, is_joinee, is_owner,
)
from sheets.services.sync import generate_fingerprint
from sheets.tasks import process_sheet_events, sync_sheet_task
from sheets.validators import apply_defaults, validate_row_data
from .google_auth import get_google_oauth_flow

logger = logging.getLogger(__name__)

User = get_user_model()


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _trigger_sheet_sync(sheet_id):
    """
    Try to dispatch a Celery sync task. Falls back to synchronous execution
    if the broker is unavailable (e.g. Redis not running locally).

    The synchronous fallback is wrapped in its own try/except so a sync
    failure doesn't crash the original HTTP request.
    """
    try:
        process_sheet_events.delay(sheet_id)
    except Exception:
        logger.info("Celery unavailable — running sync synchronously for sheet %s", sheet_id)
        try:
            process_sheet_events(sheet_id)
        except Exception:
            logger.exception("Synchronous sync fallback also failed for sheet %s", sheet_id)


def _refresh_credentials_if_needed(gc):
    """
    Refresh Google credentials if they are expired. Saves the new token to DB.
    Returns the (possibly refreshed) Credentials object.
    """
    creds = gc.get_credentials()
    if creds.expired and creds.refresh_token:
        creds.refresh(google.auth.transport.requests.Request())
        gc.save_credentials(creds)
    return creds


# ─────────────────────────────────────────────────────────────────────────────
# 1. Public / Auth Views
# ─────────────────────────────────────────────────────────────────────────────

def home(request):
    """Public landing page. Redirect authenticated users to the dashboard."""
    if request.user.is_authenticated:
        return redirect("dashboard")
    return render(request, "home.html")


def google_login(request):
    """
    Initiate the Google OAuth 2.0 authorization flow.

    Fix: If the user arrives via 127.0.0.1, the session cookie is set on that
    host. When Google redirects back to localhost:8000, the session is lost
    (different host). Force redirect to localhost before starting the flow.
    """
    if "127.0.0.1" in request.get_host():
        return redirect("http://localhost:8000/google/login/")

    flow = get_google_oauth_flow()
    auth_url, state = flow.authorization_url(
        access_type="offline",      # required to receive a refresh_token
        prompt="consent",           # forces refresh_token on every first login
        include_granted_scopes="true",
    )
    # Store OAuth state in session for CSRF verification in the callback.
    request.session["google_oauth_state"] = state
    return redirect(auth_url)


@require_POST
def logout_view(request):
    """
    Log out the current user and flush their session.

    Requires POST to prevent CSRF-triggered logouts via a crafted <img> tag
    or link. The logout button in the template must submit a form with POST.
    """
    logout(request)
    request.session.flush()  # clears Google OAuth state and all session data
    return redirect("home")


def google_callback(request):
    """
    Handle the OAuth 2.0 callback from Google.

    Security steps:
      1. Verify the OAuth state parameter matches what we stored in the session
         (prevents CSRF attacks on the OAuth flow).
      2. Exchange the authorization code for credentials.
      3. Fetch user profile via the Google userinfo API.
      4. Create or retrieve the Django user and persist credentials.
    """
    # ── Step 1: CSRF state verification ──────────────────────────────────────
    stored_state = request.session.pop("google_oauth_state", None)
    returned_state = request.GET.get("state")
    if not stored_state or stored_state != returned_state:
        logger.warning(
            "OAuth state mismatch — possible CSRF attack. stored=%s returned=%s",
            stored_state, returned_state,
        )
        return HttpResponseBadRequest(
            "Invalid OAuth state. Please try logging in again."
        )

    # ── Step 2: Exchange code for credentials ────────────────────────────────
    flow = get_google_oauth_flow()
    flow.fetch_token(authorization_response=request.build_absolute_uri())
    creds = flow.credentials

    # ── Step 3: Fetch user profile ───────────────────────────────────────────
    # Use the requests library here since creds are freshly minted (not expired).
    # A try/except guards against unexpected API failures.
    try:
        user_info_resp = requests.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {creds.token}"},
            timeout=10,
        )
        user_info_resp.raise_for_status()
        user_info = user_info_resp.json()
    except Exception:
        logger.exception("Failed to fetch Google userinfo after OAuth callback")
        return HttpResponseBadRequest(
            "Could not retrieve your Google profile. Please try logging in again."
        )

    email = user_info.get("email")
    if not email:
        return HttpResponseBadRequest(
            "Could not retrieve email from Google. Please try again."
        )

    name = user_info.get("name", "")
    picture = user_info.get("picture")

    # ── Step 4: Create/retrieve user and save credentials ────────────────────
    user, _ = User.objects.get_or_create(
        username=email,
        defaults={"email": email, "first_name": name},
    )
    login(request, user)

    gc, _ = GoogleCredentials.objects.get_or_create(user=user)
    gc.profile_picture = picture
    gc.save_credentials(creds)

    return redirect("dashboard")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Dashboard
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    """
    Main user dashboard.

    Shows three distinct sheet lists:
      - Created sheets  : owned by this user (full access)
      - Collaborating   : explicitly invited (full access, shown with badge)
      - Joined sheets   : joined via share link (own-row access only)

    Also shows the 6 most recent activity log entries relevant to this user.
    """
    user = request.user

    # ── Owned sheets (with member prefetch to avoid N+1 on member count) ─────
    owned_qs = (
        Sheet.objects
        .filter(owner=user, is_active=True)
        .annotate(row_count=Count("rows", distinct=True))
        .prefetch_related(
            "owner",
            Prefetch(
                "members",
                queryset=SheetMember.objects.filter(is_active=True).select_related("user"),
                to_attr="active_members",
            ),
        )
        .order_by("-created_at")
    )

    created_sheets = []
    total_rows = 0
    for sheet in owned_qs:
        sheet.response_count = sheet.row_count
        sheet.is_owner = True
        total_rows += sheet.row_count
        created_sheets.append(sheet)

    last_act = owned_qs.aggregate(Max("updated_at"))["updated_at__max"]

    # ── Joined sheets (JOINEE role — own rows only) ──────────────────────────
    joined_members = (
        SheetMember.objects
        .select_related("sheet", "sheet__owner")
        .filter(
            user=user,
            role=SheetMember.ROLE_JOINEE,
            is_active=True,
            sheet__is_active=True,
        )
        .exclude(sheet__owner=user)
        .order_by("-joined_at")
    )

    # ── Collaborating sheets (COLLABORATOR role — full access) ───────────────
    collaborating_members = (
        SheetMember.objects
        .select_related("sheet", "sheet__owner")
        .filter(
            user=user,
            role=SheetMember.ROLE_COLLABORATOR,
            is_active=True,
            sheet__is_active=True,
        )
        .exclude(sheet__owner=user)
        .order_by("-joined_at")
    )

    # ── Recent activity (owned sheets + actions by this user) ────────────────
    owned_ids = list(owned_qs.values_list("id", flat=True))
    recent_activity = (
        ActivityLog.objects
        .filter(Q(sheet_id__in=owned_ids) | Q(user=user))
        .select_related("user", "sheet", "user__google_credentials")
        .order_by("-created_at")[:6]
    )

    context = {
        "created_sheets": created_sheets,
        "joined_sheets": joined_members,
        "collaborating_sheets": collaborating_members,
        "total_sheets": len(created_sheets) + joined_members.count() + collaborating_members.count(),
        "total_rows": total_rows,
        "last_activity": last_act,
        "recent_activity": recent_activity,
    }

    return render(request, "dashboard.html", context)


# ─────────────────────────────────────────────────────────────────────────────
# 3. Sheet CRUD
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def create_google_sheet(request):
    """
    Create a new Google Sheet on the owner's Drive and record it in the DB.

    Accepts two column input modes (progressive enhancement):
      1. column_configs (JSON) — new typed column builder UI.
         Each item: {name, type, options, validation, default_value}
      2. columns (comma-separated string) — legacy fallback.

    Optionally add collaborators (JSON list of emails) at creation time.
    Unregistered collaborator emails return an error.
    """
    try:
        title = request.POST.get("title", "").strip()
        if not title:
            return JsonResponse({"error": "Title required"}, status=400)

        gc = GoogleCredentials.objects.get(user=request.user)
        creds = _refresh_credentials_if_needed(gc)
        service = build("sheets", "v4", credentials=creds, cache_discovery=False)

        # ── Create the Google Spreadsheet ────────────────────────────────────
        spreadsheet = service.spreadsheets().create(
            body={"properties": {"title": title}},
            fields="spreadsheetId,spreadsheetUrl",
        ).execute()
        spreadsheet_id = spreadsheet["spreadsheetId"]

        # ── Resolve column definitions ────────────────────────────────────────
        # Priority: typed JSON builder > legacy comma-separated string
        column_configs_raw = request.POST.get("column_configs", "").strip()
        columns_raw        = request.POST.get("columns", "").strip()

        sheet_columns = []    # flat list of header names for Google Sheets
        parsed_configs = []   # list of dicts for SheetColumn creation

        if column_configs_raw:
            try:
                parsed_configs = json.loads(column_configs_raw)
                if not isinstance(parsed_configs, list):
                    parsed_configs = []
            except json.JSONDecodeError:
                parsed_configs = []

            sheet_columns = [
                cfg.get("name", "").strip()
                for cfg in parsed_configs
                if cfg.get("name", "").strip()
            ]
        elif columns_raw:
            # Legacy mode: plain comma-separated column names
            sheet_columns = [c.strip() for c in columns_raw.split(",") if c.strip()]

        # Write header row to Google Sheet
        if sheet_columns:
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range="A1",
                valueInputOption="RAW",
                body={"values": [sheet_columns]},
            ).execute()

        # ── Create DB Sheet record ────────────────────────────────────────────
        sheet = Sheet.objects.create(
            owner=request.user,
            name=title,
            google_sheet_id=spreadsheet_id,
            google_url=spreadsheet["spreadsheetUrl"],
            columns=sheet_columns,
        )

        # ── Create SheetColumn records (typed columns only) ───────────────────
        if parsed_configs:
            col_objs = []
            for idx, cfg in enumerate(parsed_configs):
                col_name = cfg.get("name", "").strip()
                if not col_name:
                    continue
                col_objs.append(SheetColumn(
                    sheet       = sheet,
                    column_name = col_name,
                    column_type = cfg.get("type", SheetColumn.COLUMN_TYPE_TEXT),
                    position    = idx,
                    options     = cfg.get("options") or [],
                    validation  = cfg.get("validation") or {},
                    default_value = cfg.get("default_value", ""),
                ))
            if col_objs:
                SheetColumn.objects.bulk_create(col_objs)

        # ── Add collaborators (optional, owner only at creation) ───────────────
        collaborators_raw = request.POST.get("collaborators", "").strip()
        collab_errors = []
        if collaborators_raw:
            try:
                collab_emails = json.loads(collaborators_raw)
                if not isinstance(collab_emails, list):
                    collab_emails = []
            except json.JSONDecodeError:
                collab_emails = []

            drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)

            for email in collab_emails:
                email = email.strip().lower()
                if not email:
                    continue

                collab_user = User.objects.filter(email=email).first()
                if not collab_user:
                    collab_errors.append(
                        f"{email} is not registered — ask them to sign in to LinkSheet first."
                    )
                    continue

                if collab_user == request.user:
                    collab_errors.append(f"{email}: you cannot add yourself as a collaborator.")
                    continue

                # Share on Google Drive
                try:
                    drive_service.permissions().create(
                        fileId=spreadsheet_id,
                        body={"type": "user", "role": "writer", "emailAddress": email},
                        fields="id",
                        sendNotificationEmail=False,
                    ).execute()
                except Exception:
                    logger.warning("Drive share failed for %s on sheet %s", email, sheet.id)

                SheetMember.objects.update_or_create(
                    sheet=sheet,
                    user=collab_user,
                    defaults={"role": SheetMember.ROLE_COLLABORATOR, "is_active": True},
                )

        response_data = {
            "success": True,
            "sheet_id": sheet.id,
            "sheet_url": sheet.google_url,
            "share_link": request.build_absolute_uri(f"/join/{sheet.share_token}/"),
            "name": sheet.name,
        }
        if collab_errors:
            response_data["collab_warnings"] = collab_errors

        return JsonResponse(response_data)

    except GoogleCredentials.DoesNotExist:
        return JsonResponse(
            {"error": "Google account not connected. Please log in again."},
            status=403,
        )
    except Exception as exc:
        logger.exception("Failed to create Google Sheet for user %s", request.user.username)
        if any(kw in str(exc).lower() for kw in ("insufficient", "scope", "forbidden")):
            return JsonResponse(
                {"error": "Your Google permissions have changed. Please log out and log back in to re-authorize."},
                status=403,
            )
        return JsonResponse(
            {"error": "Failed to create sheet. Please try again or re-login."},
            status=500,
        )


@login_required
@require_POST
def delete_sheet(request, sheet_id):
    """
    Hard-delete a sheet from both the DB and Google Drive.
    Only the OWNER may delete a sheet.

    Drive deletion failure is logged but does not block the DB deletion —
    the owner can manually delete the orphaned file from Drive if needed.
    """
    try:
        sheet = get_object_or_404(Sheet, id=sheet_id, owner=request.user)
        google_sheet_id = sheet.google_sheet_id

        # Best-effort: delete from Google Drive (may fail if already deleted or creds expired).
        try:
            google_creds = GoogleCredentials.objects.get(user=request.user)
            creds = _refresh_credentials_if_needed(google_creds)
            drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
            drive_service.files().delete(fileId=google_sheet_id).execute()
        except Exception:
            logger.warning(
                "Google Drive delete failed for sheet %s — DB record will still be removed.",
                sheet_id,
            )

        sheet.delete()
        return JsonResponse({"success": True})

    except Exception:
        logger.exception("Failed to delete sheet %s", sheet_id)
        return JsonResponse(
            {"success": False, "error": "Failed to delete the sheet. Please try again."},
            status=400,
        )


@login_required
@require_GET
def download_sheet(request, sheet_id):
    """
    Download sheet data from Google Sheets as Excel (.xlsx) or CSV.
    Only OWNER and COLLABORATOR may download — JOINEE cannot.
    """
    try:
        sheet = get_object_or_404(Sheet, id=sheet_id)

        # Only OWNER and COLLABORATOR may download full sheet data.
        if not can_download_sheet(sheet, request.user):
            return HttpResponse(
                "Forbidden: only the sheet owner or collaborators may download.",
                status=403,
            )

        format_type = request.GET.get("format", "xlsx").lower()

        # Use OWNER's credentials — the file lives on their Drive.
        google_creds = GoogleCredentials.objects.get(user=sheet.owner)
        creds = _refresh_credentials_if_needed(google_creds)

        sheets_service = build("sheets", "v4", credentials=creds, cache_discovery=False)
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=sheet.google_sheet_id,
            range="A:Z",
        ).execute()

        values = result.get("values", [])

        if not values:
            return HttpResponse("No data found in sheet.", status=404)

        if format_type == "csv":
            return _download_as_csv(sheet, values)
        return _download_as_excel(sheet, values)

    except GoogleCredentials.DoesNotExist:
        return HttpResponse(
            "Sheet owner credentials not found. Please ask the owner to reconnect their Google account.",
            status=403,
        )
    except Exception:
        logger.exception("Download failed for sheet %s by user %s", sheet_id, request.user.username)
        return HttpResponse("Download failed. Please try again.", status=500)


def _download_as_csv(sheet, data):
    """Generate and return a CSV HttpResponse from a list of row lists."""
    response = HttpResponse(content_type="text/csv")
    filename = f"{sheet.name}_{timezone.now().strftime('%Y%m%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    for row in data:
        writer.writerow(row)
    return response


def _download_as_excel(sheet, data):
    """Generate and return a styled Excel (.xlsx) HttpResponse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet.name[:31]  # Excel sheet name limit is 31 chars

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for row_num, row_data in enumerate(data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit column widths (capped at 50 chars to avoid overly wide columns).
    for column_cells in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value),
            default=0,
        )
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{sheet.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# 4. Row CRUD
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def add_row(request, sheet_id):
    """
    Add a new data row to a sheet.
    Any member (OWNER, COLLABORATOR, JOINEE) may add rows.
    Applies column defaults then validates before saving.
    """
    sheet = get_object_or_404(Sheet, id=sheet_id)

    if not can_access_sheet(sheet, request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    data = json.loads(request.body or "{}")

    # Apply default values before validation (so defaults go through validation too)
    data = apply_defaults(sheet, data)

    # Server-side validation
    errors = validate_row_data(sheet, data)
    if errors:
        return JsonResponse({"error": "Validation failed", "fields": errors}, status=400)

    with transaction.atomic():
        row = SheetRow.objects.create(
            sheet=sheet,
            user=request.user,
            data=data,
            sheet_row_number=None,
        )
        SheetSyncEvent.objects.create(
            sheet=sheet,
            row=row,
            action="create",
            payload=data,
        )

    _trigger_sheet_sync(sheet.id)
    return JsonResponse({"id": row.id, "data": row.data, "pending": True}, status=201)


@login_required
@require_http_methods(["PUT"])
def update_row(request, sheet_id, row_id):
    """
    Update an existing row.
    OWNER/COLLABORATOR can edit any row. JOINEE can only edit their own.
    Validates updated data against column rules (excluding self from unique checks).
    """
    sheet = get_object_or_404(Sheet, id=sheet_id)

    if not can_access_sheet(sheet, request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    row = get_object_or_404(SheetRow, id=row_id, sheet=sheet)

    if not can_modify_row(sheet, request.user, row):
        return JsonResponse(
            {"error": "Forbidden: you can only edit your own rows."},
            status=403,
        )

    if row.sheet_row_number is None:
        return JsonResponse(
            {"error": "Row is still syncing. Please try again in a moment."},
            status=409,
        )

    data = json.loads(request.body)

    # Validate — exclude self row from unique checks so an unchanged unique
    # value doesn't falsely trigger a duplicate error
    errors = validate_row_data(sheet, data, exclude_row_id=row.id)
    if errors:
        return JsonResponse({"error": "Validation failed", "fields": errors}, status=400)

    with transaction.atomic():
        row.data = data
        row.save(update_fields=["data", "updated_at"])
        SheetSyncEvent.objects.create(
            sheet=sheet,
            row=row,
            action="update",
            payload=data,
        )

    _trigger_sheet_sync(sheet.id)
    return JsonResponse({"id": row.id, "data": row.data})


@login_required
@require_http_methods(["DELETE"])
def delete_row(request, row_id):
    """
    Delete a data row.
    OWNER/COLLABORATOR can delete any row. JOINEE can only delete their own.
    """
    row = get_object_or_404(SheetRow, id=row_id)
    sheet = row.sheet

    if not can_access_sheet(sheet, request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    if not can_delete_row(sheet, request.user, row):
        return JsonResponse(
            {"error": "Forbidden: you can only delete your own rows."},
            status=403,
        )

    with transaction.atomic():
        SheetSyncEvent.objects.create(
            sheet=sheet,
            action="delete",
            row_number=row.sheet_row_number,
            payload={},
        )
        row.delete()

    _trigger_sheet_sync(sheet.id)
    return JsonResponse({"success": True})


# ─────────────────────────────────────────────────────────────────────────────
# 5. Grid Data (AJAX)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def sheet_grid_data(request, sheet_id):
    """
    Return JSON data for the sheet grid view.

    Resolves the user's role in ONE DB query via get_role(), then applies
    row-level filtering accordingly. JOINEE only sees their own rows.

    Now also returns column_configs so the frontend can render type-aware
    inputs in the form drawer.
    """
    sheet = get_object_or_404(Sheet, id=sheet_id)

    role = get_role(sheet, request.user)
    if role is None:
        return JsonResponse({"error": "Forbidden"}, status=403)

    headers = sheet.columns
    if not headers:
        gc = GoogleCredentials.objects.filter(user=sheet.owner).first()
        if gc:
            creds = gc.get_credentials()
            service = build("sheets", "v4", credentials=creds, cache_discovery=False)
            result = service.spreadsheets().values().get(
                spreadsheetId=sheet.google_sheet_id,
                range="1:1",
            ).execute()
            headers = result.get("values", [[]])[0]
            sheet.columns = headers
            sheet.save(update_fields=["columns"])

    rows = get_visible_rows(sheet, request.user)
    column_configs = sheet.get_column_configs()

    return JsonResponse({
        "title": sheet.name,
        "columns": headers,
        "column_configs": column_configs,
        "role": role,
        "can_see_all": role in ("owner", "collaborator"),
        "rows": [
            {"id": r.id, "data": r.data, "is_own": r.user_id == request.user.pk}
            for r in rows
        ],
    })


# ─────────────────────────────────────────────────────────────────────────────
# 6. Membership
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def join_sheet(request, token):
    """
    Join a sheet via a share link. Always creates JOINEE role — no exceptions.

    Safeguards:
    - Owner cannot join their own sheet.
    - Existing collaborators are redirected without downgrading their role.
    - Re-joining an inactive membership re-activates it without changing role.
    """
    sheet = get_object_or_404(Sheet, share_token=token, is_active=True)

    # Prevent owner from joining their own sheet.
    if sheet.owner == request.user:
        messages.info(request, "You are the owner of this sheet.")
        return redirect("dashboard")

    # Prevent collaborators from accidentally getting downgraded to joinee.
    if is_collaborator(sheet, request.user):
        messages.info(request, f"You are already a collaborator on: {sheet.name}")
        return redirect("dashboard")

    # Create or re-activate JOINEE membership.
    member, created = SheetMember.objects.get_or_create(
        sheet=sheet,
        user=request.user,
        defaults={"role": SheetMember.ROLE_JOINEE},
    )

    if not created:
        if member.role != SheetMember.ROLE_JOINEE:
            # Safety guard: never silently downgrade collaborator → joinee.
            messages.info(request, f"You already have access to: {sheet.name}")
            return redirect("dashboard")
        member.is_active = True
        member.save(update_fields=["is_active"])
        messages.info(request, f"You are already a member of: {sheet.name}")
    else:
        messages.success(request, f"You have joined '{sheet.name}'. You can add and view your own data.")

    return redirect("dashboard")


@login_required
@require_POST
def add_collaborator(request, sheet_id):
    """
    Invite a user to the sheet as a collaborator (full access).
    OWNER only — collaborators cannot invite other collaborators.
    """
    # Only OWNER and existing COLLABORATORs can manage the collaborator list.
    sheet = get_object_or_404(
        Sheet,
        Q(id=sheet_id) & (Q(owner=request.user) | Q(members__user=request.user, members__role="collaborator")),
        is_active=True,
    )

    data = json.loads(request.body)
    email = data.get("email")
    if not email:
        return JsonResponse({"error": "Email is required"}, status=400)

    collab_user = User.objects.filter(email=email).first()
    if not collab_user:
        return JsonResponse(
            {"error": f"No user found with email '{email}'. They must sign up to LinkSheet first."},
            status=404,
        )

    if collab_user == request.user:
        return JsonResponse({"error": "You cannot add yourself as a collaborator."}, status=400)

    gc = GoogleCredentials.objects.filter(user=request.user).first()
    if not gc:
        return JsonResponse({"error": "No Google credentials found."}, status=400)

    creds = _refresh_credentials_if_needed(gc)

    try:
        drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
        drive_service.permissions().create(
            fileId=sheet.google_sheet_id,
            body={"type": "user", "role": "writer", "emailAddress": email},
            fields="id",
            sendNotificationEmail=False,
        ).execute()
    except Exception:
        logger.exception("Failed to share Google sheet with %s", email)
        return JsonResponse(
            {"error": "Failed to share via Google Drive. You may need full Google Drive permissions."},
            status=500,
        )

    SheetMember.objects.update_or_create(
        sheet=sheet,
        user=collab_user,
        defaults={"role": "collaborator", "is_active": True},
    )
    return JsonResponse({"success": True})


@login_required
@require_POST
def remove_collaborator(request, sheet_id):
    """
    Remove a collaborator from a sheet.
    OWNER only. Also revokes their Google Drive permission.
    """
    sheet = get_object_or_404(Sheet, id=sheet_id, owner=request.user)

    data = json.loads(request.body)
    email = data.get("email")
    if not email:
        return JsonResponse({"error": "Email is required"}, status=400)

    collab_user = User.objects.filter(email=email).first()
    if not collab_user:
        return JsonResponse({"error": "User not found."}, status=404)

    member = SheetMember.objects.filter(sheet=sheet, user=collab_user, role="collaborator").first()
    if not member:
        return JsonResponse({"error": "User is not a collaborator."}, status=400)

    gc = GoogleCredentials.objects.filter(user=request.user).first()
    if gc:
        creds = _refresh_credentials_if_needed(gc)
        try:
            drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
            permissions = drive_service.permissions().list(
                fileId=sheet.google_sheet_id,
                fields="permissions(id, emailAddress)",
            ).execute()

            for perm in permissions.get("permissions", []):
                if perm.get("emailAddress") == email:
                    drive_service.permissions().delete(
                        fileId=sheet.google_sheet_id,
                        permissionId=perm.get("id"),
                    ).execute()
                    break
        except Exception:
            logger.exception("Failed to remove Drive permission for %s", email)

    member.delete()
    return JsonResponse({"success": True})


# ─────────────────────────────────────────────────────────────────────────────
# 7. Sync
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def trigger_sync(request, sheet_id):
    """
    Manually trigger a Google Sheets sync for a specific sheet.
    Only OWNER and COLLABORATOR may trigger syncs.
    """
    sheet = get_object_or_404(Sheet, id=sheet_id)

    # Use the central permission helper instead of duplicating the role check.
    if not can_trigger_sync(sheet, request.user):
        return JsonResponse({"error": "Forbidden"}, status=403)

    if not sheet.is_syncing:
        sync_sheet_task.delay(sheet.id)

    return JsonResponse({"status": "syncing"})


# ─────────────────────────────────────────────────────────────────────────────
# 8. Activity Page
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def activity_page(request):
    """
    Full paginated activity log for the current user.

    Shows all activity on sheets the user owns or is a member of,
    plus any activity the user performed themselves. Supports:
    - Category filter  (?filter=created|updated|deleted)
    - Live search      (?q=...)
    - Pagination       (?page=N)
    """
    user = request.user

    # Collect all sheet IDs this user has any stake in.
    owned_ids = list(Sheet.objects.filter(owner=user).values_list("id", flat=True))
    member_ids = list(
        SheetMember.objects.filter(user=user, is_active=True).values_list("sheet_id", flat=True)
    )
    all_sheet_ids = list(set(owned_ids + member_ids))

    qs = (
        ActivityLog.objects
        .filter(Q(sheet_id__in=all_sheet_ids) | Q(user=user))
        .select_related("user", "sheet", "user__google_credentials")
        .order_by("-created_at")
    )

    # ── Category filter ───────────────────────────────────────────────────────
    action_filter = request.GET.get("filter", "all")
    ACTION_MAP = {
        "created": [ActivityLog.ACTION_SHEET_CREATED],
        "updated": [ActivityLog.ACTION_ROW_ADDED, ActivityLog.ACTION_ROW_UPDATED],
        "deleted": [ActivityLog.ACTION_ROW_DELETED],
    }
    if action_filter in ACTION_MAP:
        qs = qs.filter(action__in=ACTION_MAP[action_filter])

    # ── Search filter ─────────────────────────────────────────────────────────
    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            Q(sheet__name__icontains=search) | Q(user__username__icontains=search)
        )

    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "activity.html", {
        "page_obj": page_obj,
        "total": paginator.count,
        "action_filter": action_filter,
        "search": search,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 9. AJAX Helpers
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_GET
def get_created_sheets(request):
    """
    AJAX endpoint returning JSON data for the created-sheets table.
    Only responds to XMLHttpRequest calls (checked via header).
    """
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return JsonResponse({"error": "AJAX only"}, status=400)

    sheets = Sheet.objects.filter(
        owner=request.user,
        is_active=True,
    ).order_by("-created_at")

    return JsonResponse({
        "sheets": [
            {
                "name": sheet.name,
                "created_at": sheet.created_at.strftime("%b %d, %Y"),
                "response_count": sheet.response_count,
                "google_url": sheet.google_url,
            }
            for sheet in sheets
        ]
    })


# ─────────────────────────────────────────────────────────────────────────────
# 10. File Upload (Cloudinary)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def upload_file(request):
    """
    Upload a file to Cloudinary (free tier) for use in a file-type column.

    Returns the secure Cloudinary URL which is then stored as the cell value
    in the SheetRow data JSON.

    Limits:
      - Max file size: CLOUDINARY_MAX_FILE_SIZE_MB (default 10 MB)
      - Only authenticated users may upload

    POST body: multipart/form-data with field 'file'
    Response:  {url: "https://res.cloudinary.com/..."}
    """
    import cloudinary.uploader
    from django.conf import settings
    import logging

    logger = logging.getLogger(__name__)

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"error": "No file provided."}, status=400)

    max_bytes = getattr(settings, "CLOUDINARY_MAX_FILE_SIZE_MB", 10) * 1024 * 1024
    if uploaded_file.size > max_bytes:
        mb = getattr(settings, "CLOUDINARY_MAX_FILE_SIZE_MB", 10)
        return JsonResponse(
            {"error": f"File too large. Maximum size is {mb} MB."},
            status=400,
        )

    # Sanitise filename — use first 80 chars, strip path components
    import os
    safe_name = os.path.basename(uploaded_file.name)[:80]

    try:
        result = cloudinary.uploader.upload(
            uploaded_file,
            public_id=f"linksheet/{request.user.id}/{safe_name}",
            overwrite=True,
            resource_type="auto",   # handles images, PDFs, docs etc.
        )
        secure_url = result.get("secure_url", "")
        if not secure_url:
            raise ValueError("Cloudinary returned no URL")

        return JsonResponse({"url": secure_url})

    except Exception:
        logger.exception(
            "Cloudinary upload failed for user %s, file %s",
            request.user.username, safe_name
        )
        return JsonResponse(
            {"error": "File upload failed. Please check your Cloudinary credentials or try again."},
            status=500,
        )
